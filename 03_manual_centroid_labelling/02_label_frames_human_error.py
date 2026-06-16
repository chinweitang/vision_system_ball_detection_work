# 02_label_frames_human_error.py
# Collect 2 extra independent labelling repeats on a stratified sample of frames
# to quantify human noise on centroid and diameter measurements.
#
# HOW THE NOISE FLOOR IS CALCULATED
# -----------------------------------
# Each sampled frame is labelled 3 times (repeat 1 = original CSV, repeats 2 & 3
# collected fresh here with no annotation overlays shown, in randomised order).
#
# Per frame, we compute the sample standard deviation (ddof=1) across the 3 repeats:
#   std_cx_i   = stdev([r1_cx,   r2_cx,   r3_cx])   — x-centroid scatter
#   std_cy_i   = stdev([r1_cy,   r2_cy,   r3_cy])   — y-centroid scatter
#   std_2d_i   = sqrt(std_cx_i² + std_cy_i²)         — combined 2-D centroid scatter
#   std_diam_i = stdev([r1_diam, r2_diam, r3_diam])  — diameter scatter
#
# The summary metrics are the mean of these per-frame values across all N sampled frames:
#   noise_cx          = mean(std_cx_i)    [px, 1σ] — avg x-centroid bounce between repeats
#   noise_cy          = mean(std_cy_i)    [px, 1σ] — avg y-centroid bounce between repeats
#   noise_2d_centroid = mean(std_2d_i)    [px, 1σ] — avg 2-D centroid position noise
#   noise_diameter    = mean(std_diam_i)  [px, 1σ] — avg diameter noise
#
# Interpretation: noise_2d_centroid = 1.2 px means the centroid estimate varies by ~1.2 px
# (1 standard deviation) between independent labellings of the same frame. This is the
# human noise floor — any measured algorithm error smaller than this is within labelling noise.
#
# Requires:  pip install openpyxl
#
# Usage:
#   python 02_label_frames_human_error.py [--source CSV] [--folder FOLDER]
#                                          [--output XLSX] [--stride N] [--pad N]
#
# Keys:
#   s / Enter        save and advance to next frame
#   Delete           clear live clicks (redo current frame)
#   ← →              navigate within the current repeat's frame queue
#   z / 0            reset zoom to fit-screen
#   q / Esc          quit (progress is saved)
# Mouse:
#   left-click       place click 1, then click 2
#   scroll wheel     zoom in / out centred on cursor
#   right-click drag pan
#
# Output (.xlsx) has two sheets:
#   "labels"  — one row per sampled frame, columns r1_*/r2_*/r3_*
#   "summary" — noise statistics (written when all 3 repeats are complete)
#
# NOTE: close the .xlsx in Excel before running, otherwise the save will fail.

import argparse
import csv
import math
import random
import re
import statistics
from pathlib import Path

import cv2
import numpy as np
import openpyxl

# ---- default paths ----
HERE        = Path(__file__).resolve().parent
SESSION     = HERE.parent / "data" / "2026-06-01_Dyson_library_test"
DEFAULT_SRC = SESSION / "tuning" / "02_moving" / "flight_01_towards_leg_labels.csv"
DEFAULT_IN  = SESSION / "moving" / "flight_01" / "flight_01_towards_leg"
DEFAULT_OUT = SESSION / "tuning" / "02_moving" / "flight_01_towards_leg_labels_human_error.xlsx"

MAX_DISPLAY_H = 900
ZOOM_STEP     = 1.25
ZOOM_MAX      = 10.0

KEY_LEFT_WIN  = 2424832
KEY_RIGHT_WIN = 2555904
KEY_LEFT_LIN  = 65361
KEY_RIGHT_LIN = 65363
KEY_DEL_WIN   = 3014656
KEY_DEL_ASCII = 127

LABEL_FIELDS = [
    "frame_number",
    "r1_click1_x", "r1_click1_y", "r1_click2_x", "r1_click2_y",
    "r1_centroid_x", "r1_centroid_y", "r1_diameter_px",
    "r2_click1_x", "r2_click1_y", "r2_click2_x", "r2_click2_y",
    "r2_centroid_x", "r2_centroid_y", "r2_diameter_px",
    "r3_click1_x", "r3_click1_y", "r3_click2_x", "r3_click2_y",
    "r3_centroid_x", "r3_centroid_y", "r3_diameter_px",
]


# ---- data I/O ---------------------------------------------------------------

def frame_num(path: Path) -> int:
    m = re.search(r"(\d+)", path.stem)
    return int(m.group(1)) if m else 0


def load_source_csv(path: Path) -> list:
    """Return ball-labelled rows from the original labels CSV, sorted by frame_number."""
    rows = []
    with open(path, newline="") as f:
        for row in csv.DictReader(f):
            if row.get("centroid_x"):
                rows.append(row)
    return sorted(rows, key=lambda r: int(r["frame_number"]))


def load_xlsx(path: Path) -> dict:
    """Return {frame_number(int): row_dict} from the 'labels' sheet, or empty dict."""
    if not path.exists():
        return {}
    wb = openpyxl.load_workbook(path)
    if "labels" not in wb.sheetnames:
        return {}
    ws      = wb["labels"]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    out     = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        d  = {h: ("" if v is None else str(v)) for h, v in zip(headers, row)}
        fn = int(d["frame_number"])
        out[fn] = d
    return out


def save_xlsx(path: Path, labels: dict, summary=None) -> None:
    """Fully rewrite .xlsx from in-memory labels dict; optionally add summary sheet."""
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "labels"
    ws.append(LABEL_FIELDS)
    for fn in sorted(labels):
        row = labels[fn]
        ws.append([row.get(f, "") for f in LABEL_FIELDS])

    if summary:
        ws2 = wb.create_sheet("summary")
        ws2.append(["metric", "value", "unit"])
        for name, val in summary.items():
            ws2.append([name, round(val, 4), "px"])

    wb.save(path)


# ---- repeat helpers ---------------------------------------------------------

def repeat_done(row: dict, r: int) -> bool:
    return bool(row.get(f"r{r}_centroid_x"))


def determine_repeat(labels: dict, sample_fns: list):
    """Return 2, 3, or None (all done)."""
    if all(repeat_done(labels.get(fn, {}), 3) for fn in sample_fns):
        return None
    if all(repeat_done(labels.get(fn, {}), 2) for fn in sample_fns):
        return 3
    return 2


def get_repeat_order(sample_fns: list, repeat: int) -> list:
    """Deterministic shuffle keyed by repeat number — same order on resume."""
    order = list(sample_fns)
    random.Random(repeat).shuffle(order)
    return order


# ---- noise calculation ------------------------------------------------------

def compute_noise(labels: dict, sample_fns: list) -> dict:
    """Per-frame std dev across 3 repeats, averaged across all frames."""
    std_cx, std_cy, std_diam, std_2d = [], [], [], []
    for fn in sample_fns:
        row = labels.get(fn, {})
        try:
            cxs  = [float(row[f"r{r}_centroid_x"])  for r in (1, 2, 3)]
            cys  = [float(row[f"r{r}_centroid_y"])  for r in (1, 2, 3)]
            dims = [float(row[f"r{r}_diameter_px"]) for r in (1, 2, 3)]
        except (ValueError, KeyError):
            continue
        s_cx   = statistics.stdev(cxs)
        s_cy   = statistics.stdev(cys)
        s_diam = statistics.stdev(dims)
        std_cx.append(s_cx)
        std_cy.append(s_cy)
        std_diam.append(s_diam)
        std_2d.append(math.hypot(s_cx, s_cy))

    n = len(std_cx)
    if n == 0:
        return {}
    return {
        "noise_cx":          sum(std_cx)   / n,
        "noise_cy":          sum(std_cy)   / n,
        "noise_2d_centroid": sum(std_2d)   / n,
        "noise_diameter":    sum(std_diam) / n,
    }


# ---- drawing (no stored overlays — canvas always blank + live clicks only) --

def _to_pad(x, y, pad: int):
    return int(round(x)) + pad, int(round(y)) + pad


def _crosshair(canvas, cx: int, cy: int, color, size: int = 8) -> None:
    cv2.line(canvas, (cx - size, cy), (cx + size, cy), color, 1)
    cv2.line(canvas, (cx, cy - size), (cx, cy + size), color, 1)


def draw_live_clicks(canvas: np.ndarray, clicks: list, pad: int) -> None:
    """Draw click dots and (after 2nd click) centroid cross + diameter circle."""
    for x, y in clicks:
        cv2.circle(canvas, _to_pad(x, y, pad), 4, (255, 255, 255), -1)
    if len(clicks) == 2:
        c1, c2 = clicks
        cx = (c1[0] + c2[0]) / 2.0
        cy = (c1[1] + c2[1]) / 2.0
        r  = max(1, int(math.hypot(c2[0] - c1[0], c2[1] - c1[1]) / 2))
        pcx, pcy = _to_pad(cx, cy, pad)
        _crosshair(canvas, pcx, pcy, (0, 255, 255))
        cv2.circle(canvas, (pcx, pcy), r, (0, 255, 0), 1)


def build_canvas(img_gray: np.ndarray, pad: int, clicks: list) -> np.ndarray:
    """Padded canvas with live clicks only — no stored annotation ever shown."""
    padded = cv2.copyMakeBorder(img_gray, pad, pad, pad, pad, cv2.BORDER_CONSTANT, value=0)
    canvas = cv2.cvtColor(padded, cv2.COLOR_GRAY2BGR)
    if clicks:
        draw_live_clicks(canvas, clicks, pad)
    return canvas


def set_title(win: str, pos: int, total: int, fn: int,
              repeat: int, saved: bool, zoom_val: float) -> None:
    status = "SAVED" if saved else "pending"
    cv2.setWindowTitle(win,
        f"[{pos+1}/{total}] frame_{fn:03d}  r{repeat}  {status}  {zoom_val:.1f}x | "
        "[s/Enter]=save  [Del]=redo  [← →]=nav  [z]=reset-zoom  [q/Esc]=quit")


# ---- main -------------------------------------------------------------------

def main():
    ap = argparse.ArgumentParser(
        description="Collect 2 extra labelling repeats to measure human noise floor.")
    ap.add_argument("--source", type=Path, default=DEFAULT_SRC,
                    help="Original labels CSV (repeat 1 source).")
    ap.add_argument("--folder", type=Path, default=DEFAULT_IN,
                    help="Folder of PNG frames to display.")
    ap.add_argument("--output", type=Path, default=DEFAULT_OUT,
                    help="Output .xlsx file (created if absent).")
    ap.add_argument("--stride", type=int, default=5,
                    help="Sample every Nth labelled frame (default 5).")
    ap.add_argument("--pad",    type=int, default=50,
                    help="Border padding in pixels (default 50).")
    args = ap.parse_args()

    folder   = args.folder.resolve()
    pad      = args.pad
    src_path = args.source.resolve()
    out_path = args.output.resolve()

    # ---- load source CSV and build stratified sample ----
    src_rows   = load_source_csv(src_path)
    sampled    = src_rows[::args.stride]
    sample_fns = [int(r["frame_number"]) for r in sampled]
    print(f"Source: {src_path.name}  ({len(src_rows)} labelled frames)")
    print(f"Sample: every {args.stride}th → {len(sample_fns)} frames: {sample_fns}")

    # Build lookup: frame_number → PNG path
    fn_to_path = {frame_num(p): p for p in folder.glob("*.png")}
    missing    = [fn for fn in sample_fns if fn not in fn_to_path]
    if missing:
        print(f"WARNING: {len(missing)} sampled frame(s) have no PNG: {missing}")
        sample_fns = [fn for fn in sample_fns if fn in fn_to_path]
    if not sample_fns:
        print("No frames to label. Exiting.")
        return

    # ---- load existing output or start fresh ----
    labels = load_xlsx(out_path)

    # Populate r1 from source CSV for any sampled frame not yet in labels
    r1_added = 0
    for r in sampled:
        fn = int(r["frame_number"])
        if fn not in sample_fns:
            continue
        if fn not in labels:
            labels[fn] = {f: "" for f in LABEL_FIELDS}
            labels[fn]["frame_number"] = fn
        row = labels[fn]
        if not row.get("r1_centroid_x"):
            row["r1_click1_x"]    = r.get("click1_x", "")
            row["r1_click1_y"]    = r.get("click1_y", "")
            row["r1_click2_x"]    = r.get("click2_x", "")
            row["r1_click2_y"]    = r.get("click2_y", "")
            row["r1_centroid_x"]  = r.get("centroid_x", "")
            row["r1_centroid_y"]  = r.get("centroid_y", "")
            row["r1_diameter_px"] = r.get("diameter_px", "")
            r1_added += 1
    if r1_added:
        save_xlsx(out_path, labels)
        print(f"Populated r1 for {r1_added} frame(s) from source CSV → {out_path.name}")

    # ---- shared mutable display state ----
    clicks     = []
    img_cache  = [None]
    cur_fn     = [None]
    scale      = [1.0]
    fit_dims   = [None]
    zoom       = [1.0]
    pan_x      = [0.0]
    pan_y      = [0.0]
    is_panning = [False]
    drag_start = [None]
    # ctx is read by on_mouse for title updates during zoom; updated before each frame load
    ctx = {"pos": 0, "total": 0, "repeat": 2}

    # Compute display scale from first frame (all frames same resolution)
    _peek    = cv2.imread(str(fn_to_path[sample_fns[0]]), cv2.IMREAD_GRAYSCALE)
    _ph, _pw = _peek.shape[0] + 2*pad, _peek.shape[1] + 2*pad
    scale[0]    = min(1.0, MAX_DISPLAY_H / _ph)
    fit_dims[0] = (int(round(_pw * scale[0])), int(round(_ph * scale[0])))

    WIN = "Human Error Labeller"
    cv2.namedWindow(WIN, cv2.WINDOW_AUTOSIZE)

    def clamp_pan():
        fw, fh = fit_dims[0]
        pan_x[0] = max(0.0, min(pan_x[0], fw - fw / zoom[0]))
        pan_y[0] = max(0.0, min(pan_y[0], fh - fh / zoom[0]))

    def refresh():
        canvas = build_canvas(img_cache[0], pad, clicks)
        if scale[0] < 1.0:
            fit = cv2.resize(canvas, None, fx=scale[0], fy=scale[0],
                             interpolation=cv2.INTER_AREA)
        else:
            fit = canvas
        fw, fh = fit_dims[0]
        if zoom[0] > 1.0:
            vw = fw / zoom[0]
            vh = fh / zoom[0]
            x0 = int(round(max(0.0, min(pan_x[0], fw - vw))))
            y0 = int(round(max(0.0, min(pan_y[0], fh - vh))))
            x1 = min(x0 + int(round(vw)), fw)
            y1 = min(y0 + int(round(vh)), fh)
            disp = cv2.resize(fit[y0:y1, x0:x1], (fw, fh),
                              interpolation=cv2.INTER_NEAREST)
        else:
            disp = fit
        cv2.imshow(WIN, disp)

    def load_fn(fn: int) -> None:
        clicks.clear()
        cur_fn[0]    = fn
        img_cache[0] = cv2.imread(str(fn_to_path[fn]), cv2.IMREAD_GRAYSCALE)
        refresh()
        saved = repeat_done(labels.get(fn, {}), ctx["repeat"])
        set_title(WIN, ctx["pos"], ctx["total"], fn, ctx["repeat"], saved, zoom[0])

    def on_mouse(event, x, y, flags, param):
        if event == cv2.EVENT_LBUTTONDOWN and len(clicks) < 2:
            fit_x  = x / zoom[0] + pan_x[0]
            fit_y  = y / zoom[0] + pan_y[0]
            clicks.append((int(round(fit_x / scale[0])) - pad,
                           int(round(fit_y / scale[0])) - pad))
            refresh()

        elif event == cv2.EVENT_MOUSEWHEEL:
            factor = ZOOM_STEP if flags > 0 else 1.0 / ZOOM_STEP
            new_z  = max(1.0, min(zoom[0] * factor, ZOOM_MAX))
            cx = x / zoom[0] + pan_x[0]
            cy = y / zoom[0] + pan_y[0]
            zoom[0]  = new_z
            pan_x[0] = cx - x / zoom[0]
            pan_y[0] = cy - y / zoom[0]
            clamp_pan()
            refresh()
            saved = repeat_done(labels.get(cur_fn[0], {}), ctx["repeat"])
            set_title(WIN, ctx["pos"], ctx["total"], cur_fn[0],
                      ctx["repeat"], saved, zoom[0])

        elif event == cv2.EVENT_RBUTTONDOWN:
            is_panning[0] = True
            drag_start[0] = (x, y, pan_x[0], pan_y[0])

        elif event == cv2.EVENT_MOUSEMOVE and is_panning[0]:
            sx, sy, px0, py0 = drag_start[0]
            pan_x[0] = px0 - (x - sx) / zoom[0]
            pan_y[0] = py0 - (y - sy) / zoom[0]
            clamp_pan()
            refresh()

        elif event == cv2.EVENT_RBUTTONUP:
            is_panning[0] = False

    cv2.setMouseCallback(WIN, on_mouse)

    # ---- outer repeat loop (repeat 2 → repeat 3 → done) ----
    quit_flag = False
    while not quit_flag:
        cur_repeat = determine_repeat(labels, sample_fns)

        if cur_repeat is None:
            # All repeats complete — compute and save noise statistics
            print("\nAll repeats complete. Computing noise statistics...")
            noise = compute_noise(labels, sample_fns)
            save_xlsx(out_path, labels, summary=noise)
            print("\n--- Human labelling noise floor ---")
            for name, val in noise.items():
                print(f"  {name:<22}  {val:.4f} px")
            print(f"\nResults saved to: {out_path}")
            break

        # Build the ordered queue for this repeat (only unsaved frames, in shuffle order)
        full_order = get_repeat_order(sample_fns, cur_repeat)
        order      = [fn for fn in full_order
                      if not repeat_done(labels.get(fn, {}), cur_repeat)]
        if not order:
            break  # safety — shouldn't happen if determine_repeat is correct

        print(f"\nRepeat {cur_repeat}: {len(order)} frame(s) remaining "
              f"(randomised order, seed={cur_repeat}) ...")

        ctx["repeat"] = cur_repeat
        ctx["total"]  = len(order)
        ctx["pos"]    = 0
        load_fn(order[ctx["pos"]])

        # ---- inner frame loop for one repeat ----
        while ctx["pos"] < len(order) and not quit_flag:
            key = cv2.waitKeyEx(50)
            if key == -1:
                continue

            fn = cur_fn[0]

            if key in (ord("q"), 27):                        # quit
                quit_flag = True

            elif key in (KEY_LEFT_WIN, KEY_LEFT_LIN):        # navigate ←
                if ctx["pos"] > 0:
                    ctx["pos"] -= 1
                    load_fn(order[ctx["pos"]])

            elif key in (KEY_RIGHT_WIN, KEY_RIGHT_LIN):      # navigate →
                if ctx["pos"] < len(order) - 1:
                    ctx["pos"] += 1
                    load_fn(order[ctx["pos"]])

            elif key in (ord("z"), ord("0")):                # reset zoom
                zoom[0] = 1.0; pan_x[0] = 0.0; pan_y[0] = 0.0
                refresh()
                saved = repeat_done(labels.get(fn, {}), cur_repeat)
                set_title(WIN, ctx["pos"], ctx["total"], fn,
                          cur_repeat, saved, zoom[0])

            elif key in (KEY_DEL_WIN, KEY_DEL_ASCII):        # redo
                clicks.clear()
                refresh()

            elif key in (ord("s"), 13):                      # save
                if len(clicks) != 2:
                    print(f"frame {fn:03d}: need 2 clicks first (have {len(clicks)})")
                    continue

                c1   = (float(clicks[0][0]), float(clicks[0][1]))
                c2   = (float(clicks[1][0]), float(clicks[1][1]))
                cx   = (c1[0] + c2[0]) / 2.0
                cy   = (c1[1] + c2[1]) / 2.0
                diam = math.hypot(c2[0] - c1[0], c2[1] - c1[1])
                r    = cur_repeat

                row = labels.setdefault(fn, {f: "" for f in LABEL_FIELDS})
                row["frame_number"]       = fn
                row[f"r{r}_click1_x"]    = f"{c1[0]:.1f}"
                row[f"r{r}_click1_y"]    = f"{c1[1]:.1f}"
                row[f"r{r}_click2_x"]    = f"{c2[0]:.1f}"
                row[f"r{r}_click2_y"]    = f"{c2[1]:.1f}"
                row[f"r{r}_centroid_x"]  = f"{cx:.4f}"
                row[f"r{r}_centroid_y"]  = f"{cy:.4f}"
                row[f"r{r}_diameter_px"] = f"{diam:.4f}"
                save_xlsx(out_path, labels)
                print(f"frame {fn:03d}  r{r}: diameter={diam:.1f}px")

                ctx["pos"] += 1
                if ctx["pos"] < len(order):
                    load_fn(order[ctx["pos"]])

        if not quit_flag and cur_repeat == 2:
            print("\nRepeat 2 complete. Starting repeat 3 ...")
        # Outer loop continues to determine_repeat() → repeat 3 or done

    cv2.destroyAllWindows()
    if quit_flag:
        print("Quit early — progress saved to:", out_path)


if __name__ == "__main__":
    main()
